import io
import json
import os
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from enum import Enum

import httpx
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import Column, DateTime, Float, Integer, String, Text, create_engine
from sqlalchemy.orm import DeclarativeBase, Session, sessionmaker

DATABASE_URL = os.getenv(
    "DATABASE_URL", "postgresql://invite:invite123@localhost:5432/order_db"
)
CART_SERVICE_URL = os.getenv("CART_SERVICE_URL", "http://localhost:8002")
ADDRESS_SERVICE_URL = os.getenv("ADDRESS_SERVICE_URL", "http://localhost:8004")
PAYMENT_SERVICE_URL = os.getenv("PAYMENT_SERVICE_URL", "http://localhost:8005")
TRACKING_SERVICE_URL = os.getenv("TRACKING_SERVICE_URL", "http://localhost:8006")

engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)


class Base(DeclarativeBase):
    pass


class OrderStatus(str, Enum):
    PENDING = "pending"
    CONFIRMED = "confirmed"
    FAILED = "failed"


class Order(Base):
    __tablename__ = "orders"

    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(String(50), unique=True, index=True, nullable=False)
    session_id = Column(String(100), nullable=False)
    customer_email = Column(String(200), nullable=False)
    subtotal = Column(Float, nullable=False)
    shipping_cost = Column(Float, default=0)
    total_amount = Column(Float, nullable=False)
    status = Column(String(20), default=OrderStatus.PENDING.value)
    payment_id = Column(String(50), nullable=True)
    items_json = Column(Text, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)


class DeliveryAddress(Base):
    __tablename__ = "delivery_addresses"

    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(String(50), index=True, nullable=False)
    recipient_name = Column(String(100), nullable=False)
    street = Column(String(200), nullable=False)
    city = Column(String(100), nullable=False)
    state = Column(String(100), nullable=False)
    postal_code = Column(String(10), nullable=False)
    country = Column(String(100), nullable=False)
    phone = Column(String(20), nullable=True)
    is_valid = Column(Integer, default=1)
    validation_issues = Column(Text, nullable=True)


class AddressInput(BaseModel):
    recipient_name: str = Field(min_length=2)
    street: str = Field(min_length=5)
    city: str = Field(min_length=2)
    state: str = Field(min_length=2)
    postal_code: str = Field(min_length=4)
    country: str = Field(min_length=2)
    phone: str | None = None


class PaymentDetails(BaseModel):
    method: str = Field(pattern="^(card|upi|netbanking|wallet)$")
    card_number: str | None = None
    card_holder: str | None = None


class CheckoutRequest(BaseModel):
    session_id: str
    customer_email: str
    addresses: list[AddressInput] = Field(min_length=1)
    payment: PaymentDetails


class OrderResponse(BaseModel):
    order_id: str
    status: str
    subtotal: float
    shipping_cost: float
    total_amount: float
    payment_id: str | None
    payment_status: str | None
    address_count: int
    message: str


class ParsedAddress(BaseModel):
    row_number: int
    recipient_name: str
    street: str
    city: str
    state: str
    postal_code: str
    country: str
    phone: str | None
    is_valid: bool
    issues: list[str]


@asynccontextmanager
async def lifespan(app: FastAPI):
    Base.metadata.create_all(bind=engine)
    yield


app = FastAPI(title="Order Service", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


async def validate_addresses(addresses: list[AddressInput]) -> list[dict]:
    async with httpx.AsyncClient() as client:
        response = await client.post(
            f"{ADDRESS_SERVICE_URL}/validate/bulk",
            json={"addresses": [a.model_dump() for a in addresses]},
            timeout=30,
        )
        response.raise_for_status()
        return response.json()["results"]


def parse_excel_addresses(file_content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file must have a header row and at least one data row")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = {"name", "street", "city", "state", "postal_code", "country"}
    header_map = {}
    aliases = {
        "name": ["name", "recipient_name", "recipient", "full name"],
        "street": ["street", "address", "street_address", "address line"],
        "city": ["city", "town"],
        "state": ["state", "province", "region"],
        "postal_code": ["postal_code", "pincode", "zip", "zipcode", "pin"],
        "country": ["country", "nation"],
        "phone": ["phone", "mobile", "contact", "phone number"],
    }

    for field, names in aliases.items():
        for i, h in enumerate(headers):
            if h in names:
                header_map[field] = i
                break

    missing = required - set(header_map.keys())
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel missing required columns: {', '.join(missing)}. "
            f"Expected columns like: name, street, city, state, postal_code, country",
        )

    addresses = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        addr = {
            "row_number": row_num,
            "recipient_name": str(row[header_map["name"]] or "").strip(),
            "street": str(row[header_map["street"]] or "").strip(),
            "city": str(row[header_map["city"]] or "").strip(),
            "state": str(row[header_map["state"]] or "").strip(),
            "postal_code": str(row[header_map["postal_code"]] or "").strip(),
            "country": str(row[header_map["country"]] or "").strip(),
            "phone": str(row[header_map["phone"]]).strip() if "phone" in header_map and row[header_map["phone"]] else None,
        }
        if addr["recipient_name"] and addr["street"]:
            addresses.append(addr)

    if not addresses:
        raise HTTPException(status_code=400, detail="No valid address rows found in Excel")

    return addresses


@app.get("/health")
def health():
    return {"status": "healthy", "service": "order-service"}


@app.post("/orders/parse-excel")
async def parse_and_validate_excel(file: UploadFile = File(...)):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    content = await file.read()
    parsed = parse_excel_addresses(content)

    address_inputs = [
        AddressInput(
            recipient_name=a["recipient_name"],
            street=a["street"],
            city=a["city"],
            state=a["state"],
            postal_code=a["postal_code"],
            country=a["country"],
            phone=a.get("phone"),
        )
        for a in parsed
    ]
    validation_results = await validate_addresses(address_inputs)

    results = []
    for addr, validation in zip(parsed, validation_results):
        results.append(
            ParsedAddress(
                row_number=addr["row_number"],
                recipient_name=addr["recipient_name"],
                street=addr["street"],
                city=addr["city"],
                state=addr["state"],
                postal_code=addr["postal_code"],
                country=addr["country"],
                phone=addr.get("phone"),
                is_valid=validation["is_valid"],
                issues=validation["issues"],
            )
        )

    return {
        "total": len(results),
        "valid_count": sum(1 for r in results if r.is_valid),
        "invalid_count": sum(1 for r in results if not r.is_valid),
        "addresses": results,
    }


@app.post("/orders/validate-addresses")
async def validate_address_list(addresses: list[AddressInput]):
    results = await validate_addresses(addresses)
    return {
        "total": len(results),
        "valid_count": sum(1 for r in results if r["is_valid"]),
        "invalid_count": sum(1 for r in results if not r["is_valid"]),
        "results": [
            {
                "address": addresses[i].model_dump(),
                "is_valid": r["is_valid"],
                "confidence_score": r["confidence_score"],
                "issues": r["issues"],
                "normalized_address": r.get("normalized_address"),
            }
            for i, r in enumerate(results)
        ],
    }


@app.post("/orders/checkout", response_model=OrderResponse, status_code=201)
async def checkout(payload: CheckoutRequest, db: Session = next(get_db())):
    async with httpx.AsyncClient() as client:
        cart_resp = await client.get(f"{CART_SERVICE_URL}/cart/{payload.session_id}")
        if cart_resp.status_code != 200:
            raise HTTPException(status_code=400, detail="Could not fetch cart")
        cart = cart_resp.json()

    if not cart["items"]:
        raise HTTPException(status_code=400, detail="Cart is empty")

    validation_results = await validate_addresses(payload.addresses)
    invalid = [
        (i, r) for i, r in enumerate(validation_results) if not r["is_valid"]
    ]
    if invalid:
        details = [
            f"Address {i + 1}: {', '.join(r['issues'])}" for i, r in invalid
        ]
        raise HTTPException(
            status_code=422,
            detail={"message": "Some addresses failed validation", "errors": details},
        )

    subtotal = cart["subtotal"]
    shipping_cost = round(len(payload.addresses) * 2.99, 2)
    total_amount = round(subtotal + shipping_cost, 2)
    order_id = f"ORD-{uuid.uuid4().hex[:12].upper()}"

    order = Order(
        order_id=order_id,
        session_id=payload.session_id,
        customer_email=payload.customer_email,
        subtotal=subtotal,
        shipping_cost=shipping_cost,
        total_amount=total_amount,
        items_json=json.dumps(cart["items"]),
    )
    db.add(order)

    for addr, validation in zip(payload.addresses, validation_results):
        normalized = validation.get("normalized_address") or addr.model_dump()
        db.add(
            DeliveryAddress(
                order_id=order_id,
                recipient_name=normalized["recipient_name"],
                street=normalized["street"],
                city=normalized["city"],
                state=normalized["state"],
                postal_code=normalized["postal_code"],
                country=normalized["country"],
                phone=normalized.get("phone"),
                is_valid=1,
                validation_issues=json.dumps(validation["issues"]) if validation["issues"] else None,
            )
        )

    db.commit()

    async with httpx.AsyncClient() as client:
        payment_resp = await client.post(
            f"{PAYMENT_SERVICE_URL}/payments",
            json={
                "order_id": order_id,
                "amount": total_amount,
                "method": payload.payment.method,
                "card_number": payload.payment.card_number,
                "card_holder": payload.payment.card_holder,
            },
            timeout=30,
        )
        payment_resp.raise_for_status()
        payment = payment_resp.json()

        await client.post(
            f"{TRACKING_SERVICE_URL}/tracking",
            json={
                "order_id": order_id,
                "status": "placed",
                "message": "Order placed successfully",
                "location": "Online Store",
            },
        )

        if payment["status"] == "completed":
            order.status = OrderStatus.CONFIRMED.value
            order.payment_id = payment["payment_id"]
            await client.post(
                f"{TRACKING_SERVICE_URL}/tracking",
                json={
                    "order_id": order_id,
                    "status": "payment_confirmed",
                    "message": "Payment confirmed",
                    "location": "Payment Gateway",
                },
            )
            await client.delete(f"{CART_SERVICE_URL}/cart/{payload.session_id}")
        else:
            order.status = OrderStatus.FAILED.value

        db.commit()

    return OrderResponse(
        order_id=order_id,
        status=order.status,
        subtotal=subtotal,
        shipping_cost=shipping_cost,
        total_amount=total_amount,
        payment_id=payment.get("payment_id"),
        payment_status=payment.get("status"),
        address_count=len(payload.addresses),
        message=payment.get("message", "Order processed"),
    )


@app.get("/orders/{order_id}")
def get_order(order_id: str, db: Session = next(get_db())):
    order = db.query(Order).filter(Order.order_id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    addresses = db.query(DeliveryAddress).filter(DeliveryAddress.order_id == order_id).all()
    return {
        "order_id": order.order_id,
        "status": order.status,
        "customer_email": order.customer_email,
        "subtotal": order.subtotal,
        "shipping_cost": order.shipping_cost,
        "total_amount": order.total_amount,
        "payment_id": order.payment_id,
        "items": json.loads(order.items_json),
        "addresses": [
            {
                "recipient_name": a.recipient_name,
                "street": a.street,
                "city": a.city,
                "state": a.state,
                "postal_code": a.postal_code,
                "country": a.country,
                "phone": a.phone,
                "is_valid": bool(a.is_valid),
            }
            for a in addresses
        ],
        "created_at": order.created_at.isoformat(),
    }


@app.get("/orders")
def list_orders(email: str | None = None, db: Session = next(get_db())):
    query = db.query(Order)
    if email:
        query = query.filter(Order.customer_email == email)
    orders = query.order_by(Order.created_at.desc()).limit(50).all()
    return [
        {
            "order_id": o.order_id,
            "status": o.status,
            "total_amount": o.total_amount,
            "customer_email": o.customer_email,
            "created_at": o.created_at.isoformat(),
        }
        for o in orders
    ]
